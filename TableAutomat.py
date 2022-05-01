import openpyxl as opx
import openpyxl.styles as styles
import openpyxl.utils as ou

def fill():

    people = []
    chair_num = 6
    table_num = 100
    dict = {'name' : 'C','surname' : 'D','isItPair' : 'G', 'role' : 'A', 'field' : 'T'}
    room = []

    redFill = styles.PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')
    input_file = opx.load_workbook("input.xlsx")
    output_file = opx.load_workbook("output.xlsx")
    wsInput = input_file.active
    wsOutput = output_file.active
    input_tab = []
    output_tab = []
    column_date = []

    class Table:
        def __init__(self, type, chair_num):
            self.type = type
            self.chairFree = chair_num
            self.guests = []
            for j in range(chair_num):
                self.guests.append(None)


    class Person:
        def __init__(self, name, surname, isItPair, role, field):
            self.name = name
            self.surname = surname
            self.isItPair = isItPair
            self.role = role
            self.field = field


    #table generator
    for i in range(table_num):
        room.append(Table(None, 6))

    print(room)
    print(len(room))
    print(len(room[0].guests))

    y1 = int(input("od ktorego wiersza zaczac"))
    y2 = int(input("do ktorego wiersza czytac"))

    for row in range(y1,y2+1):
        name = wsOutput[dict['name']+str(row)]
        surname = wsOutput[dict['surname'] + str(row)]
        isItPair = wsOutput[dict['isItPair'] + str(row)]
        role = wsOutput[dict['role'] + str(row)]
        field = wsOutput[dict['field'] + str(row)]
        people.append(Person(name,surname,isItPair,role,field))

    print(people[0].surname)





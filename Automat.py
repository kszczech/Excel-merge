import openpyxl as opx
import openpyxl.styles as styles
import openpyxl.utils as ou

def fill():
    redFill = styles.PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')
    input_file = opx.load_workbook("input.xlsx")
    output_file = opx.load_workbook("output.xlsx")
    wsInput = input_file.active
    wsOutput = output_file.active
    input_tab = []
    output_tab = []
    column_date = []
    column_in =  ['B','C','D','E','F','G','H','I','J','L',"M","N"] #wpisz tu z inputa kolumny do przetworzenia
    column_out = ['C','D','E','F','T','U','V','G','H','X',"Y","Z"] #wpisz tu z outputu kolumny ktore odpowiadaja kolumna z inputu
    for i in range(len(column_in)):
        line = []
        line.append(column_in[i])
        line.append(column_out[i])
        column_date.append(line)

    #print(column_date)
    print("pytanie do inputu")
    y1 = int(input("od ktorego wiersza zaczac"))
    y2 = int(input("do ktorego wiersza czytac"))

    for row in range(y1, y2+1):
        line = []
        for col in column_date:
            char = col[0] + str(row)
            line.append(wsInput[char].value)
        input_tab.append(line)
    #print(input_tab)

    print()
    print("pytania do outputu")
    y1 = int(input("od ktorego wiersza zaczac"))
    y2 = int(input("do ktorego wiersza czytac"))

    for row in range(y1, y2+1):
        line = []
        for col in column_date:
            char = col[1] + str(row)
            line.append(wsOutput[char].value)
        output_tab.append(line)
    #print(output_tab)

    #print(output_tab[1][0])

    for out in range(len(output_tab)):
        for inp in range(len(input_tab)):
            if output_tab[out][0] == input_tab[inp][0] and output_tab[out][1] == input_tab[inp][1]:
                for i in range(len(output_tab[out])):
                    if output_tab[out][i] == None:
                        output_tab[out][i] = input_tab[inp][i]
                        char = "A" + str(inp + 2)
                        wsInput[char].fill = redFill


#print(output_tab)

    for i in range(len(output_tab)):
        for j in range(len(column_date)):
            char = column_date[j][1] + str(i + y1)
            wsOutput[char].value = output_tab[i][j]

    nazwa = input("jak nazwac gotowy plik")
    output_file.save(nazwa + '.xlsx')
    input_file.save(nazwa + "_input"+".xlsx")
